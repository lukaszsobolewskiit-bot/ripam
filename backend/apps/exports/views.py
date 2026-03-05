import io
import json
import tempfile
from datetime import date

from django.core.management import call_command
from django.http import HttpResponse
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.ipam.models import VLAN, Host, Subnet, Tunnel
from apps.ipam.models import PatchPanel, PatchPanelPort, PatchPanelConnection
from apps.ipam.models import SubscriberBox, SubscriberBoxPort, SubscriberBoxConnection
from apps.ipam.models import Rack, RackUnit
from apps.ipam.models import PortConnection
from apps.projects.models import Project, Site


class IsAdmin(IsAuthenticated):
    def has_permission(self, request, view):
        return super().has_permission(request, view) and getattr(request.user, "is_admin", False)


class DataExportView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        output = io.StringIO()
        call_command(
            "dumpdata",
            "projects", "ipam", "accounts", "audit", "templates",
            format="json",
            indent=2,
            stdout=output,
        )
        content = output.getvalue()
        response = HttpResponse(content, content_type="application/json")
        response["Content-Disposition"] = f'attachment; filename="sobnet-backup-{date.today()}.json"'
        return response


class DataImportView(APIView):
    permission_classes = [IsAdmin]
    parser_classes = [MultiPartParser]

    def post(self, request):
        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"detail": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            content = uploaded.read().decode("utf-8")
            data = json.loads(content)
            if not isinstance(data, list):
                raise ValueError("Expected a JSON array")
        except (json.JSONDecodeError, ValueError, UnicodeDecodeError) as e:
            return Response({"detail": f"Invalid JSON file: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        # Validate that all entries have model field from allowed apps
        allowed_apps = {"projects.", "ipam.", "accounts.", "audit.", "templates."}
        for entry in data:
            model = entry.get("model", "")
            if not any(model.startswith(app) for app in allowed_apps):
                return Response(
                    {"detail": f"Unexpected model '{model}'. Only projects, ipam, accounts, audit data allowed."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        replace = request.query_params.get("replace") == "true"

        if replace:
            call_command("flush", "--no-input")

        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=True) as f:
            f.write(content)
            f.flush()
            try:
                call_command("loaddata", f.name)
            except Exception as e:
                return Response({"detail": f"Import failed: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        return Response({"detail": f"Imported {len(data)} objects.", "count": len(data)})


class ProjectExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        try:
            import openpyxl
        except ImportError:
            return HttpResponse("openpyxl not installed", status=500)

        project = Project.objects.get(pk=project_id)
        wb = openpyxl.Workbook()

        # Sites sheet
        ws = wb.active
        ws.title = "Sites"
        ws.append(["Name", "Address", "Latitude", "Longitude"])
        for site in project.sites.all():
            ws.append([site.name, site.address, str(site.latitude or ""), str(site.longitude or "")])

        # VLANs sheet
        ws2 = wb.create_sheet("VLANs")
        ws2.append(["Site", "VLAN ID", "Name", "Purpose"])
        for vlan in VLAN.objects.filter(site__project=project).select_related("site"):
            ws2.append([vlan.site.name, vlan.vlan_id, vlan.name, vlan.purpose])

        # Subnets sheet
        ws3 = wb.create_sheet("Subnets")
        ws3.append(["Site", "VLAN", "Network", "Gateway", "Description"])
        for subnet in Subnet.objects.filter(project=project).select_related("site", "vlan"):
            site_name = subnet.site.name
            vlan_label = f"VLAN {subnet.vlan.vlan_id}" if subnet.vlan else "(standalone)"
            ws3.append([
                site_name, vlan_label,
                str(subnet.network), str(subnet.gateway or ""), subnet.description,
            ])

        # Hosts sheet
        ws4 = wb.create_sheet("Hosts")
        ws4.append(["Site", "VLAN", "Subnet", "IP", "Hostname", "MAC", "Device Type"])
        for host in Host.objects.filter(
            subnet__project=project
        ).select_related("subnet__site", "subnet__vlan"):
            site_name = host.subnet.site.name
            vlan_label = f"VLAN {host.subnet.vlan.vlan_id}" if host.subnet.vlan else "(standalone)"
            ws4.append([
                site_name,
                vlan_label,
                str(host.subnet.network),
                str(host.ip_address),
                host.hostname,
                host.mac_address,
                host.device_type,
            ])

        # Tunnels sheet
        ws5 = wb.create_sheet("Tunnels")
        ws5.append(["Name", "Type", "Subnet", "Site A", "IP A", "Site B", "IP B", "Enabled"])
        for tunnel in Tunnel.objects.filter(project=project).select_related("site_a", "site_b"):
            ws5.append([
                tunnel.name, tunnel.tunnel_type, str(tunnel.tunnel_subnet),
                tunnel.site_a.name, str(tunnel.ip_a),
                tunnel.site_b.name, str(tunnel.ip_b),
                "Yes" if tunnel.enabled else "No",
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{project.name}.xlsx"'
        return response


class ProjectPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        project = Project.objects.get(pk=project_id)
        sites = project.sites.all()

        html = f"""
        <html>
        <head><style>
            body {{ font-family: sans-serif; font-size: 12px; }}
            h1 {{ color: #1a1a2e; }}
            table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 6px 8px; text-align: left; }}
            th {{ background-color: #f0f0f0; }}
        </style></head>
        <body>
        <h1>{project.name}</h1>
        <p>{project.description}</p>
        <p>Supernet: {project.supernet or 'N/A'}</p>
        """

        for site in sites:
            html += f"<h2>{site.name}</h2><p>{site.address}</p>"
            vlans = site.vlans.prefetch_related("subnets__hosts")
            for vlan in vlans:
                html += f"<h3>VLAN {vlan.vlan_id} - {vlan.name}</h3>"
                for subnet in vlan.subnets.all():
                    html += f"<h4>{subnet.network}</h4>"
                    html += "<table><tr><th>IP</th><th>Hostname</th><th>Type</th></tr>"
                    for host in subnet.hosts.all():
                        html += f"<tr><td>{host.ip_address}</td><td>{host.hostname}</td><td>{host.device_type}</td></tr>"
                    html += "</table>"

            # Standalone subnets for this site
            standalone = Subnet.objects.filter(site=site, vlan__isnull=True).prefetch_related("hosts")
            if standalone.exists():
                html += "<h3>Standalone Subnets</h3>"
                for subnet in standalone:
                    html += f"<h4>{subnet.network}</h4>"
                    html += "<table><tr><th>IP</th><th>Hostname</th><th>Type</th></tr>"
                    for host in subnet.hosts.all():
                        html += f"<tr><td>{host.ip_address}</td><td>{host.hostname}</td><td>{host.device_type}</td></tr>"
                    html += "</table>"

        html += "</body></html>"

        try:
            from weasyprint import HTML
            pdf = HTML(string=html).write_pdf()
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="{project.name}.pdf"'
            return response
        except ImportError:
            return HttpResponse(html, content_type="text/html")


# ─── Physical Export (Patch Panels + Subscriber Boxes + Port Connections) ─────

def _css():
    return """
    <style>
      body{font-family:Arial,sans-serif;font-size:11px;color:#1a1a2e;margin:20px}
      h1{font-size:18px;color:#1a1a2e;border-bottom:2px solid #4f46e5;padding-bottom:6px}
      h2{font-size:14px;color:#4f46e5;margin-top:18px;margin-bottom:4px}
      h3{font-size:12px;color:#374151;margin:10px 0 3px}
      table{border-collapse:collapse;width:100%;margin-bottom:14px;font-size:10px}
      th{background:#4f46e5;color:#fff;padding:4px 6px;text-align:left}
      td{border:1px solid #e5e7eb;padding:3px 6px}
      tr:nth-child(even) td{background:#f9fafb}
      .badge{display:inline-block;padding:1px 6px;border-radius:3px;font-size:9px;font-weight:600}
      .ok{background:#d1fae5;color:#065f46}.warn{background:#fef3c7;color:#92400e}
      .info{background:#dbeafe;color:#1e40af}.section{margin-top:24px}
      .stat{display:inline-block;margin-right:16px;font-size:10px}
    </style>"""


class PhysicalExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse("openpyxl not installed", status=500)

        project = Project.objects.get(pk=project_id)

        wb = openpyxl.Workbook()
        hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        hdr_fill  = PatternFill("solid", start_color="4F46E5")
        alt_fill  = PatternFill("solid", start_color="F9FAFB")
        thin      = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )

        def style_sheet(ws, headers, col_widths):
            ws.append(headers)
            for i, h in enumerate(headers, 1):
                c = ws.cell(1, i)
                c.font = hdr_font; c.fill = hdr_fill
                c.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]
            ws.freeze_panes = "A2"

        def style_rows(ws, start=2):
            for row_idx, row in enumerate(ws.iter_rows(min_row=start), start):
                fill = alt_fill if row_idx % 2 == 0 else None
                for cell in row:
                    cell.border = thin
                    if fill: cell.fill = fill

        # ── Sheet 1: Patch Panels ──────────────────────────────────────────
        ws1 = wb.active; ws1.title = "Patch Panels"
        style_sheet(ws1,
            ["Site", "Panel", "Media", "Ports", "Location", "Description"],
            [18, 22, 14, 7, 18, 28])
        panels = PatchPanel.objects.filter(
            site__project=project
        ).select_related("site").prefetch_related("ports")
        for p in panels:
            ws1.append([
                p.site.name if p.site else "",
                p.name, p.media_type, p.port_count,
                p.location, p.description,
            ])
        style_rows(ws1)

        # ── Sheet 2: Patch Panel Ports ────────────────────────────────────
        ws2 = wb.create_sheet("PP Ports")
        style_sheet(ws2,
            ["Site", "Panel", "Port #", "Label", "Back Media", "Connected Host", "Host Port", "Far Panel", "Far Port"],
            [16, 20, 7, 14, 12, 20, 12, 20, 10])
        for p in panels:
            for port in p.ports.all():
                info = getattr(port, "device_port_info", None) or {}
                ws2.append([
                    p.site.name if p.site else "",
                    p.name,
                    port.port_number,
                    port.label,
                    port.back_media_type,
                    info.get("host_name", "") if isinstance(info, dict) else "",
                    info.get("device_port_name", "") if isinstance(info, dict) else "",
                    info.get("far_panel_name", "") if isinstance(info, dict) else "",
                    info.get("far_panel_port_number", "") if isinstance(info, dict) else "",
                ])
        style_rows(ws2)

        # ── Sheet 3: Port Connections ─────────────────────────────────────
        ws3 = wb.create_sheet("Port Connections")
        style_sheet(ws3,
            ["Host A", "Port A", "Host B", "Port B", "Description"],
            [22, 14, 22, 14, 30])
        conns = PortConnection.objects.filter(
            host_a__subnet__project=project
        ).select_related("host_a", "host_b", "port_a", "port_b")
        for c in conns:
            ws3.append([
                c.host_a.hostname or str(c.host_a.ip_address),
                c.port_a.name if c.port_a else "",
                c.host_b.hostname or str(c.host_b.ip_address),
                c.port_b.name if c.port_b else "",
                getattr(c, "description", ""),
            ])
        style_rows(ws3)

        # ── Sheet 4: Subscriber Boxes ─────────────────────────────────────
        ws4 = wb.create_sheet("Subscriber Boxes")
        style_sheet(ws4,
            ["Site", "Box", "Type", "Location", "Trunk Ports", "Drop Ports", "Connected"],
            [16, 22, 14, 20, 11, 11, 10])
        boxes = SubscriberBox.objects.filter(
            site__project=project
        ).select_related("site").prefetch_related("ports")
        for b in boxes:
            ports = list(b.ports.all())
            trunk = sum(1 for p in ports if p.direction == "trunk")
            drop  = sum(1 for p in ports if p.direction == "drop")
            connected = sum(1 for p in ports if hasattr(p, "connection") and p.connection)
            ws4.append([
                b.site.name if b.site else "",
                b.name, b.box_type, b.location,
                trunk, drop, connected,
            ])
        style_rows(ws4)

        # ── Sheet 5: Subscriber Box Ports ────────────────────────────────
        ws5 = wb.create_sheet("Box Ports")
        style_sheet(ws5,
            ["Site", "Box", "Direction", "Port #", "Label", "Media", "Connected Panel", "Panel Port"],
            [16, 20, 10, 7, 16, 12, 22, 10])
        for b in boxes:
            for port in b.ports.all():
                conn_info = getattr(port, "connection_info", None) or {}
                ws5.append([
                    b.site.name if b.site else "",
                    b.name,
                    port.direction,
                    port.port_number,
                    port.label,
                    port.media_type,
                    conn_info.get("panel_name", "") if isinstance(conn_info, dict) else "",
                    conn_info.get("panel_port_number", "") if isinstance(conn_info, dict) else "",
                ])
        style_rows(ws5)

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{project.name}-physical.xlsx"'
        return resp


class PhysicalPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        project = Project.objects.get(pk=project_id)
        panels  = PatchPanel.objects.filter(
            site__project=project
        ).select_related("site").prefetch_related("ports")
        boxes   = SubscriberBox.objects.filter(
            site__project=project
        ).select_related("site").prefetch_related("ports")
        conns   = PortConnection.objects.filter(
            host_a__subnet__project=project
        ).select_related("host_a", "host_b", "port_a", "port_b")

        html = f"<html><head>{_css()}</head><body>"
        html += f"<h1>Physical — {project.name}</h1>"
        html += f"<p class='stat'>Panele: <b>{panels.count()}</b></p>"
        html += f"<p class='stat'>Puszki: <b>{boxes.count()}</b></p>"
        html += f"<p class='stat'>Połączenia portów: <b>{conns.count()}</b></p>"

        # Patch panels per site
        html += "<div class='section'><h2>Patch Panele</h2>"
        by_site: dict = {}
        for p in panels:
            k = p.site.name if p.site else "Nieprzypisane"
            by_site.setdefault(k, []).append(p)
        for site_name, site_panels in sorted(by_site.items()):
            html += f"<h3>{site_name}</h3>"
            html += "<table><tr><th>Panel</th><th>Media</th><th>Porty</th><th>Lokalizacja</th></tr>"
            for p in site_panels:
                used = sum(1 for pp in p.ports.all() if pp.device_port_info)
                html += f"<tr><td>{p.name}</td><td>{p.media_type}</td>"
                html += f"<td>{used}/{p.port_count}</td><td>{p.location or '—'}</td></tr>"
            html += "</table>"
        html += "</div>"

        # Subscriber boxes
        html += "<div class='section'><h2>Puszki Abonenckie</h2>"
        by_site2: dict = {}
        for b in boxes:
            k = b.site.name if b.site else "Nieprzypisane"
            by_site2.setdefault(k, []).append(b)
        for site_name, site_boxes in sorted(by_site2.items()):
            html += f"<h3>{site_name}</h3>"
            html += "<table><tr><th>Puszka</th><th>Typ</th><th>Trunk</th><th>Drop</th><th>Lokalizacja</th></tr>"
            for b in site_boxes:
                ports = list(b.ports.all())
                trunk = sum(1 for p in ports if p.direction == "trunk")
                drop  = sum(1 for p in ports if p.direction == "drop")
                html += f"<tr><td>{b.name}</td><td>{b.box_type}</td>"
                html += f"<td>{trunk}</td><td>{drop}</td><td>{b.location or '—'}</td></tr>"
            html += "</table>"
        html += "</div>"

        # Port connections
        if conns.exists():
            html += "<div class='section'><h2>Połączenia Portów</h2>"
            html += "<table><tr><th>Host A</th><th>Port A</th><th>Host B</th><th>Port B</th></tr>"
            for c in conns:
                ha = c.host_a.hostname or str(c.host_a.ip_address)
                hb = c.host_b.hostname or str(c.host_b.ip_address)
                pa = c.port_a.name if c.port_a else "—"
                pb = c.port_b.name if c.port_b else "—"
                html += f"<tr><td>{ha}</td><td>{pa}</td><td>{hb}</td><td>{pb}</td></tr>"
            html += "</table></div>"

        html += "</body></html>"

        try:
            from weasyprint import HTML
            pdf = HTML(string=html).write_pdf()
            resp = HttpResponse(pdf, content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="{project.name}-physical.pdf"'
            return resp
        except ImportError:
            return HttpResponse(html, content_type="text/html")


# ─── Racks Export ─────────────────────────────────────────────────────────────

class RacksExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse("openpyxl not installed", status=500)

        project = Project.objects.get(pk=project_id)
        racks   = Rack.objects.filter(
            site__project=project
        ).select_related("site").prefetch_related("rack_units__host", "rack_units__patch_panel")

        wb = openpyxl.Workbook()
        hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        hdr_fill = PatternFill("solid", start_color="4F46E5")
        alt_fill = PatternFill("solid", start_color="F9FAFB")
        thin     = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )

        def style_sheet(ws, headers, widths):
            ws.append(headers)
            for i, _ in enumerate(headers, 1):
                c = ws.cell(1, i)
                c.font = hdr_font; c.fill = hdr_fill
                c.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
            ws.freeze_panes = "A2"

        def style_rows(ws):
            for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
                fill = alt_fill if ri % 2 == 0 else None
                for cell in row:
                    cell.border = thin
                    if fill: cell.fill = fill

        # ── Sheet 1: Racks overview ──────────────────────────────────────
        ws1 = wb.active; ws1.title = "Szafy Rack"
        style_sheet(ws1,
            ["Site", "Szafa", "Facility ID", "Status", "Typ", "Wys. (U)", "Użyte (U)", "Wolne (U)",
             "Szer. (mm)", "Głęb. (mm)", "Nr seryjny", "Lokalizacja", "Opis"],
            [16, 20, 12, 14, 16, 8, 8, 8, 10, 10, 16, 20, 28])
        for r in racks:
            used = r.rack_units.count()
            ws1.append([
                r.site.name, r.name, r.facility_id, r.status, r.rack_type,
                r.height_u, used, r.height_u - used,
                r.width_mm, r.depth_mm, r.serial_number, r.location, r.description,
            ])
        style_rows(ws1)

        # ── Sheet 2: Rack units ──────────────────────────────────────────
        ws2 = wb.create_sheet("Jednostki Rack")
        style_sheet(ws2,
            ["Site", "Szafa", "Pozycja U", "Wys. U", "Typ", "Nazwa/Label",
             "Host IP", "Patch Panel", "Strona", "Kolor"],
            [16, 20, 10, 7, 14, 24, 16, 20, 7, 10])
        for r in racks:
            for u in r.rack_units.all().order_by("position_u"):
                name = u.label
                if u.host:
                    name = u.host.hostname or str(u.host.ip_address)
                elif u.patch_panel:
                    name = u.patch_panel.name
                host_ip = str(u.host.ip_address) if u.host else ""
                pp_name = u.patch_panel.name if u.patch_panel else ""
                ws2.append([
                    r.site.name, r.name, u.position_u, u.height_u,
                    u.item_type, name, host_ip, pp_name, u.face, u.color,
                ])
        style_rows(ws2)

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{project.name}-racks.xlsx"'
        return resp


class RacksPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        project = Project.objects.get(pk=project_id)
        racks   = Rack.objects.filter(
            site__project=project
        ).select_related("site").prefetch_related("rack_units__host", "rack_units__patch_panel")

        html = f"<html><head>{_css()}</head><body>"
        html += f"<h1>Szafy Rack — {project.name}</h1>"
        html += f"<p class='stat'>Szafy: <b>{racks.count()}</b></p>"

        by_site: dict = {}
        for r in racks:
            by_site.setdefault(r.site.name, []).append(r)

        for site_name, site_racks in sorted(by_site.items()):
            html += f"<div class='section'><h2>{site_name}</h2>"
            for rack in site_racks:
                units = list(rack.rack_units.all().order_by("position_u"))
                used  = len(units)
                pct   = round(used / rack.height_u * 100) if rack.height_u else 0
                html += f"<h3>{rack.name}"
                if rack.facility_id:
                    html += f" <small>({rack.facility_id})</small>"
                html += f"</h3>"
                html += f"<p class='stat'>Status: <b>{rack.status}</b></p>"
                html += f"<p class='stat'>Typ: <b>{rack.rack_type}</b></p>"
                html += f"<p class='stat'>Wysokość: <b>{rack.height_u}U</b></p>"
                html += f"<p class='stat'>Zajętość: <b>{used}/{rack.height_u}U ({pct}%)</b></p>"
                if rack.location:
                    html += f"<p class='stat'>Lokalizacja: <b>{rack.location}</b></p>"
                if units:
                    html += """<table>
                    <tr><th>U</th><th>Nazwa / Label</th><th>Typ</th><th>IP</th><th>Strona</th></tr>"""
                    for u in units:
                        name = u.label
                        if u.host:
                            name = u.host.hostname or str(u.host.ip_address)
                        elif u.patch_panel:
                            name = u.patch_panel.name
                        ip = str(u.host.ip_address) if u.host else "—"
                        html += f"<tr><td>U{u.position_u}</td><td>{name or '—'}</td>"
                        html += f"<td>{u.item_type}</td><td>{ip}</td><td>{u.face}</td></tr>"
                    html += "</table>"
                else:
                    html += "<p><i>Szafa pusta</i></p>"
            html += "</div>"

        html += "</body></html>"

        try:
            from weasyprint import HTML
            pdf = HTML(string=html).write_pdf()
            resp = HttpResponse(pdf, content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="{project.name}-racks.pdf"'
            return resp
        except ImportError:
            return HttpResponse(html, content_type="text/html")
